import unicodedata
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from ioformats.filerw import FileWriter, normalize
from ioformats import availableWriters,TABLE,BIBLIOGRAPHY

class XlsxWriter(FileWriter):
    def __init__(self,numbered=False,outputDir='.',multiSheetOutput=False,editMode=False):
        super().__init__(numbered,outputDir,multiSheetOutput,editMode,TABLE,BIBLIOGRAPHY)
        self.extension = '.xlsx'
        self.col = 1

    def _getopendoc(self,name=None):
        if name == None:
            return Workbook(write_only=True) # return a fresh one
        return load_workbook(name)
    
    def _savedoc(self,filename):
        self.doc.save(filename)

    def openSheet(self,sheetname,sheetType=TABLE,**kwargs):
        super().openSheet(sheetname,sheetType,**kwargs)
        if self.editMode:
            self.currentSheet = self.doc[sheetname]
        else:
            self.currentSheet = self.doc.create_sheet(sheetname)

    def startNewLine(self): 
        self._incLineCount()
        self.col = 1

    def writeTitle(self,iterable,always=False,**kwargs):
        ''' write a title line. if 'always', does it even in editing mode'''
        if always or not self.editMode:
            self.writeln(iterable)
    
    def writeln(self,iterable,insertMode=False,** kwargs):
        if self.editMode:
            self._writeln(iterable, insertMode)
        else:
            self.currentSheet.append(iterable)
        self._incLineCount()


    def append(self,element,insertMode=False,** kwargs):
        line = self.getCurrentLine()
        cell = self.currentSheet.cell(row=line,column=self.col)
        if insertMode: #copy previous line's cell value
            above = self.currentSheet.cell(row=line-1,column=self.col)
            cell._style = above._style
            cell.number_format = above.number_format
            if element == '' and type(above.value) is str and above.value.startswith('='): # need to translate formula
                cell.value = Translator(above.value, origin=above.coordinate).translate_formula(cell.coordinate)
        if element != '' and not cell.protection.locked:
            cell.value = element
        self.col += 1
      
    def _writeln(self,iterable,insertMode=False):
        line = self.getCurrentLine()
        self.col = 1
        if insertMode:
            self.currentSheet.insert_rows(line)
        for v in iterable:
            self.append(v,insertMode=insertMode)


# @deprecated("Use XlsxWriter")
# class XlsxUpdater(XlsxWriter):
#     def __init__(self, outdir):
#         super().__init__(outdir)
# 
#     def isWriteOnly(self):
#         return False
#         
#     def open(self, target):
#         super().open(target)
#         self.doc = load_workbook(target+self.extension)
# 
#     def openSheet(self,sheetname):
#         super().openSheet(sheetname)
#         self.currentSheet = self.doc[sheetname]
#     
#     def writeTitle(self,iterable,insertMode=False):
#         self.currentline += 1 # do not write again the title lines
# 
#     def writeln(self,iterable,insertMode=False):
#         line = self.currentline
#         if insertMode:
#             self.currentSheet.insert_rows(line)
#         col = 1
#         for v in iterable:
#             cell = self.currentSheet.cell(row=line,column=col)
#             if insertMode: #copy previous line's cell value
#                 above = self.currentSheet.cell(row=line-1,column=col)
#                 cell._style = above._style
#                 cell.number_format = above.number_format
#                 if v == '' and type(above.value) is str and above.value.startswith('='): # need to translate formula
#                     cell.value = Translator(above.value, origin=above.coordinate).translate_formula(cell.coordinate)
#             if v != '' and not cell.protection.locked:
#                 cell.value = v
#             col += 1
#         self.currentline += 1
#         
#     def closeSheet(self):
#         pass
#     
#     def close(self):
#         self.doc.save(self.outputDir+"/"+self.target+self.extension)

availableWriters['xlsx'] = XlsxWriter()
availableWriters['xlsx-edit'] = XlsxWriter(editMode=True,multiSheetOutput=True)
availableWriters['xlsx-multisheets'] = XlsxWriter(multiSheetOutput=True)


from openpyxl.cell import Cell

Cell.__init__.__defaults__ = (None, None, '', None)   # Change the default value for the Cell from None to `` the same way as in csv.DictReader


class DictReader(object):
    def __init__(self, f, sheet_index = 0, title_line = 1,
                 fieldnames=None, normalize_fieldnames=False, restkey=None, restval=None):
        self._fieldnames = fieldnames   # list of keys for the dict
        self.normalize_fieldnames = normalize_fieldnames # whether firld names should be normalized (converted to ascii)
        self.restkey  = restkey         # key to catch long rows
        self.restval  = restval         # default value for short rows
        self.reader   = load_workbook(f, data_only=True).worksheets[sheet_index].iter_rows(values_only=True)
        self.title_line = title_line    # location of the title line
        self.line_num = 0

    def __iter__(self):
        return self

    @property
    def fieldnames(self):
        if self._fieldnames is None:
            for l in range(1,self.title_line):
                next(self.reader) # skip all lines before the title
                self.line_num += 1
            try:
                self._fieldnames =  list(map(normalize,next(self.reader))) if self.normalize_fieldnames else next(self.reader)
                    
                self.line_num += 1
            except StopIteration:
                pass

        return self._fieldnames

    @fieldnames.setter
    def fieldnames(self, value):
        self._fieldnames = value

    def __next__(self):
        if self.line_num == 0:
            # Used only for its side effect.
            self.fieldnames

        row = next(self.reader)
        self.line_num += 1

        # unlike the basic reader, we prefer not to return blanks,
        # because we will typically wind up with a dict full of None
        # values
        while row == ():
            row = next(self.reader)

        d = dict(zip(self.fieldnames, row))
        lf = len(self.fieldnames)
        lr = len(row)

        if lf < lr:
            d[self.restkey] = row[lf:]
        elif lf > lr:
            for key in self.fieldnames[lr:]:
                d[key] = self.restval

        return d
    
    