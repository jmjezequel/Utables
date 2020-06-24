import logging
import os
import tempfile
import csv
import unittest
from datetime import datetime

from openpyxl import load_workbook

from ioformats import TABLE,TEXT,BIBLIOGRAPHY,LIST
from ioformats.docxrw import DocxWriter
from ioformats.tex import TeXWriter, TeXTableReader
from ioformats.text import TextWriter
from ioformats.writers import AbstractWriter
from ioformats.xlsx import XlsxWriter

TEST_DATA = [
    ['Test Table', 'String', 'Int', 'Float', 'Date'],
    ['Line 1', 'foo', 42, 3.14, datetime.fromisoformat('2020-01-01')],
    ['Line 2', '%èé&à_$', 666666666666, -0.14, datetime(2020, 1, 1)],
    ['Line 3', '', -1, 0, ''],
]

outdir = tempfile.gettempdir() + "/TestWriterMethods/"

class PublicationStub:
    def __init__(self, **kwargs):
        self.kw = kwargs

    def write(self,writer,citationStyle=None,**kwargs):
        """ use writer to output this publication in biblio form."""
        writer.append(self.kw.get('key'),key=True,href='http://www.irisa.fr/')
        writer.append(self.kw.get('authors'),authors=True)
        writer.append(self.kw.get('title'),title=True)
        v = self.kw.get('venue','')
        if v != '':
            writer.append(v+', ')
        p = self.kw.get('publisher','')
        if p != '':
            writer.append(p+', ')
        writer.append(str(self.kw.get('year'))+'.')

    def __iter__(self):
        for v in self.kw.values():
            yield v


PUBLIST = [
    PublicationStub(key='J20',authors="JM Jézéquel",title='Testing Publication Generator',venue='IXXX2020',year=2020),
    PublicationStub(key='DD20',authors='M. Dupont and P. Durand',title='Dummy Publication',publisher='Nobody',year=2020)
]

def gen_publist(writer: AbstractWriter, numbered=False, resetCount=True):
    writer.openSheet("publications", BIBLIOGRAPHY, numbered, resetCount)
    for pub in PUBLIST:
        writer.writeln(pub)
    writer.closeSheet()

def gen_sheet(writer: AbstractWriter, sheetType: str, label: str, numbered=False, resetCount=True, **kargs):
    writer.openSheet('test_'+sheetType+label,sheetType,numbered,resetCount,**kargs)
    writer.writeTitle(TEST_DATA[0])
    for i in range(1, len(TEST_DATA)):
        writer.writeln(TEST_DATA[i])
    writer.closeSheet()

def gen_tables(writer: AbstractWriter, sheets: int = 1):
    for i in range(1,sheets+1):
        gen_sheet(writer,TABLE, str(i))

def gen_text(writer: AbstractWriter, numbered=False, resetCount=True, **kargs):
    writer.openSheet('test_'+TEXT,TEXT,numbered,resetCount,**kargs)
    writer.writeTitle('Test title level 1',level=1)
    for i in range(1,3):
        writer.writeTitle(('Test title', 'level 2'), level=2)
        for j in range(1,3):
            writer.writeln('Lorem ipsum lorem ipsum '+str(i)+ '.'+str(j))
            writer.writeln('Italic text',italic=True)
            writer.writeln('Bold text',bold=True)
            writer.writeln('Normal text')
            writer.writeln('')
    writer.closeSheet()

def get_writers(sheetType: str):
    return filter(lambda w: w.isSupporting(sheetType), availableWriters.values())

def run_writer(w: AbstractWriter, label: str, numbered=False, resetCount=True, **kargs):
    w.setOutputDir(outdir)
    w.open(label)
    for kind in w.getSupportedTypes():
        print('Writing ',kind,' with ',label)
        if kind == BIBLIOGRAPHY:
            gen_publist(w, True)
        elif kind == TABLE:
            gen_tables(w, 2)
        elif kind == TEXT:
            gen_text(w, numbered,resetCount,**kargs)
        else:
            gen_sheet(w, kind, label, numbered,resetCount,**kargs)
    w.close()


class TestWriterMethods(unittest.TestCase):
    def setUp(self):
        os.makedirs(outdir, exist_ok=True)

    def tearDown(self):
        pass

    def check_standalone_tables(self, extension:str, number:int=1):
        for i in range(1,number+1):
            self.check_table(outdir + extension + '-' + BASETABLENAME + str(i) + '.' + extension, BASETABLENAME+str(i))
    # def test_writers(self):
    #     for k, w in availableWriters.items():
    #         run_writer(w,k)
    #

BASETABLENAME ='test_table'
class Test_CSV(TestWriterMethods):

    def test_writer(self):
        from ioformats.csvwriter import CSVwriter
        w = CSVwriter()
        run_writer(w,'csv')
        self.check_standalone_tables('csv',2)

    def check_table(self, filename, * sheetnames):
        with open(filename,'r', encoding='utf-8') as csvfile:
            r = csv.reader(csvfile, delimiter=';')
            line = 0
            for row in r:
                # only check column of type string cause read does not
                # yield back int/floats
                self.assertEqual(TEST_DATA[line][0],row[0])
                self.assertEqual(TEST_DATA[line][1],row[1])
                #self.assertEqual(TEST_DATA[line],row)
                line += 1

class Test_xlsx(TestWriterMethods):
    def test_single_sheet_writer(self):
        w = XlsxWriter()
        run_writer(w,'xlsx')
        self.check_standalone_tables('xlsx',2)

    def test_multiple_sheet_writer(self):
        w = XlsxWriter(multiSheetOutput=True)
        run_writer(w,'xlsx-multisheets')
        self.check_table(outdir+'xlsx-multisheets.xlsx',BASETABLENAME+'1',BASETABLENAME+'2')

    def check_table(self, filename, * sheetnames):
        wb = load_workbook(filename)
        for sn in sheetnames:
            ws = wb[sn]
            line = 0
            for row in ws.values:
                self.assertEqual(TEST_DATA[line],list(elem if elem is not None else '' for elem in row))
                line += 1

class Test_docx(TestWriterMethods):
    def test_single_sheet_writer(self):
        w = DocxWriter()
        run_writer(w,'docx')

    def test_multiple_sheet_writer(self):
        w = DocxWriter(multiSheetOutput=True)
        run_writer(w,'docx-article')
        # self.check_table(outdir+'xlsx-multisheets.xlsx',BASETABLENAME+'1',BASETABLENAME+'2')

    def check_table(self, filename, * sheetnames):
        wb = load_workbook(filename)
        for sn in sheetnames:
            ws = wb.get_sheet_by_name(sn)
            line = 0
            for row in ws.values:
                self.assertEqual(TEST_DATA[line],list(elem if elem is not None else '' for elem in row))
                line += 1

class Test_txt(TestWriterMethods):
    def test_single_sheet_writer(self):
        w = TextWriter()
        run_writer(w,'txt')
        self.check_standalone_tables('txt',2)

    def test_multiple_sheet_writer(self):
        w = TextWriter(multiSheetOutput=True)
        run_writer(w,'txt-article')
        # self.check_table(outdir+'xlsx-multisheets.xlsx',BASETABLENAME+'1',BASETABLENAME+'2')

    def check_table(self, filename, * sheetnames):
        with open(filename,'r', encoding='utf-8') as file:
            line = 0
            for row in file:
                self.assertEqual(' '.join(str(e) for e in TEST_DATA[line]),row.rstrip('\n'))
                line += 1


class Test_tex(TestWriterMethods):
    def test_single_sheet_writer(self):
        w = TeXWriter()
        run_writer(w,'tex')
        self.check_standalone_tables('tex',2)

    def test_multiple_sheet_writer(self):
        w = TextWriter(multiSheetOutput=True)
        run_writer(w,'tex-article')
        # self.check_table(outdir+'xlsx-multisheets.xlsx',BASETABLENAME+'1',BASETABLENAME+'2')

    def check_table(self, filename, * sheetnames):
        with open(filename,'r', encoding='utf-8') as file:
            reader = TeXTableReader(file)
            line = 0
            for row in reader:
                self.assertEqual(TEST_DATA[line],row)
                line += 1


#        self.assertEqual('foo'.upper(), 'FOO')

    # def test_isupper(self):
    #     self.assertTrue('FOO'.isupper())
    #     self.assertFalse('Foo'.isupper())
    #
    # def test_split(self):
    #     s = 'hello world'
    #     self.assertEqual(s.split(), ['hello', 'world'])
    #     # check that s.split fails when the separator is not a string
    #     with self.assertRaises(TypeError):
    #         s.split(2)


if __name__ == '__main__':
    unittest.main()